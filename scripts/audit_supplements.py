"""Validate downloaded image and spreadsheet supplementary files."""

from __future__ import annotations

import argparse
import hashlib
import json
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def inspect(path: Path) -> dict[str, object]:
    result: dict[str, object] = {
        "name": path.name,
        "bytes": path.stat().st_size,
        "sha256": sha256(path),
    }
    suffix = path.suffix.lower()
    if suffix in {".tif", ".tiff", ".png", ".jpg", ".jpeg", ".webp"}:
        with Image.open(path) as image:
            image.verify()
        with Image.open(path) as image:
            result.update(
                {
                    "kind": "image",
                    "format": image.format,
                    "width": image.width,
                    "height": image.height,
                    "frames": getattr(image, "n_frames", 1),
                }
            )
    elif suffix == ".xlsx":
        workbook = load_workbook(path, read_only=True, data_only=False)
        try:
            result.update(
                {
                    "kind": "workbook",
                    "worksheets": len(workbook.sheetnames),
                    "sheet_names": workbook.sheetnames,
                }
            )
        finally:
            workbook.close()
    else:
        result["kind"] = "unhandled"
    return result


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("root", type=Path)
    args = parser.parse_args()
    files = sorted(path for path in args.root.iterdir() if path.is_file())
    print(json.dumps([inspect(path) for path in files], indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
